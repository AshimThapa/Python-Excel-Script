import openpyxl
import time
from calendar import monthrange
from openpyxl import load_workbook

days=[]
file_no=[]
trips=[]
amount=[]
day_list=[]

month=time.strftime('%b')
year=int(time.strftime('%Y'))
month_num=int(time.strftime('%m'))
day_list=monthrange(year,month_num)[1]


m1=['M4','EASTLINK','SABL','SACL']
#function to clear items from array before each for loop.
def clearArrays():
    file_no.clear()
    trips.clear()
    amount.clear()
    return;

#this function gets called from Auto_Report.py
def CallMainPage(m,wb,logger):
    logger.info("Preparing Main Page")
    if m=='M4WCX':
        m=m1[0]
    elif m=='ELK':
        m=m1[1]
    elif m=='SAB':
        m=m1[2]
    elif m=='SAT':
        m=m1[3]

    #select the working sheet    
    ws=wb['M5 ON '+m]
    #clear arrays
    days.clear()
    clearArrays()
    #call for loop for the days in month
    for i in range(1,day_list+1):
        #convert date 1 to 01 and so on using Zfill(2) function
        #if i<10:
        file_no.append("'"+str(i).zfill(2)+month+"'!D22")#list of filenumber of a month
        trips.append("'"+str(i).zfill(2)+month+"'!E24")#list of number of trips for a month
        amount.append("'"+str(i).zfill(2)+month+"'!F24")#list of trip amounts for a month
        days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))#Date fromat in dd-Mon-yy            
        #else:
           # file_no.append("'"+str(i)+month+"'!D22")
           # trips.append("'"+str(i)+month+"'!E24")
           # amount.append("'"+str(i)+month+"'!F24")
           # days.append(str(i)+"-"+month+"-"+time.strftime('%y'))
        

        #THE CELL LOCATION AND CORRESPONDING DATA    
        ws['B'+str(10+i)]=days[i-1]
        ws['C'+str(10+i)]='='+file_no[i-1]
        ws['D'+str(10+i)]='='+trips[i-1]
        ws['E'+str(10+i)]='='+amount[i-1]
    logger.info("111 Data inserted")

    clearArrays()
    for i in range(1,day_list+1):        
        
        if i<10:
            file_no.append("'"+str(i).zfill(2)+month+"'!D28")
            trips.append("'"+str(i).zfill(2)+month+"'!E30")
            amount.append("'"+str(i).zfill(2)+month+"'!F30")
            #days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))
        else:
            file_no.append("'"+str(i)+month+"'!D28")
            trips.append("'"+str(i)+month+"'!E30")
            amount.append("'"+str(i)+month+"'!F30")
            #days.append(str(i)+"-"+month+"-"+time.strftime('%y'))

        ws['O'+str(10+i)]=days[i-1]
        ws['P'+str(10+i)]='='+file_no[i-1]
        ws['Q'+str(10+i)]='='+trips[i-1]
        ws['R'+str(10+i)]='='+amount[i-1]
    logger.info("108 Data inserted")

    clearArrays()
    for i in range(1,day_list+1):       
        
        if i<10:
            file_no.append("'"+str(i).zfill(2)+month+"'!D34")
            trips.append("'"+str(i).zfill(2)+month+"'!E35")
            amount.append("'"+str(i).zfill(2)+month+"'!F35")
            #days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))
        else:
            file_no.append("'"+str(i)+month+"'!D34")
            trips.append("'"+str(i)+month+"'!E35")
            amount.append("'"+str(i)+month+"'!F35")
            #days.append(str(i)+"-"+month+"-"+time.strftime('%y))

        ws['C'+str(48+i)]='='+file_no[i-1]
        ws['D'+str(48+i)]='='+trips[i-1]
        ws['E'+str(48+i)]='='+amount[i-1]
    logger.info("file 33 111 Data inserted")

    clearArrays()
    for i in range(1,day_list+1):
        
               
        if i<10:
            file_no.append("'"+str(i).zfill(2)+month+"'!D39")
            trips.append("'"+str(i).zfill(2)+month+"'!E40")
            amount.append("'"+str(i).zfill(2)+month+"'!F40")
            #days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))
        else:
            file_no.append("'"+str(i)+month+"'!D39")
            trips.append("'"+str(i)+month+"'!E40")
            amount.append("'"+str(i)+month+"'!F40")
            #days.append(str(i)+"-"+month+"-"+time.strftime('%y))

        ws['C'+str(81+i)]='='+file_no[i-1]
        ws['D'+str(81+i)]='='+trips[i-1]
        ws['E'+str(81+i)]='='+amount[i-1]
    logger.info("file 33 108 Data inserted")

    clearArrays()

    for i in range(1,day_list+1):            
        if i<10:
            file_no.append("'"+str(i).zfill(2)+month+"'!H54")
            trips.append("'"+str(i).zfill(2)+month+"'!I54")
            amount.append("'"+str(i).zfill(2)+month+"'!J54")
            #days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))
        else:
            file_no.append("'"+str(i)+month+"'!H54")
            trips.append("'"+str(i)+month+"'!I54")
            amount.append("'"+str(i)+month+"'!J54")
            #days.append(str(i)+"-"+month+"-"+time.strftime('%y))

        ws['N'+str(48+i)]='='+file_no[i-1]
        ws['O'+str(48+i)]='='+trips[i-1]
        ws['P'+str(48+i)]='='+amount[i-1]
    logger.info("file 33 108 Data inserted")

    for i in range(1,day_list+1):
        if i<10:
            file_no.append("'"+str(i).zfill(2)+month+"'!H55")
            trips.append("'"+str(i).zfill(2)+month+"'!I55")
            amount.append("'"+str(i).zfill(2)+month+"'!J55")
            #days.append(str(i).zfill(2)+"-"+month+"-"+time.strftime('%y'))
        else:
            file_no.append("'"+str(i)+month+"'!H55")
            trips.append("'"+str(i)+month+"'!I55")
            amount.append("'"+str(i)+month+"'!J55")
            #days.append(str(i)+"-"+month+"-"+time.strftime('%y))

        ws['N'+str(81+i)]='='+file_no[i-1]
        ws['O'+str(81+i)]='='+trips[i-1]
        ws['P'+str(81+i)]='='+amount[i-1]
    logger.info("file 33 108 Data inserted")

    return;    
