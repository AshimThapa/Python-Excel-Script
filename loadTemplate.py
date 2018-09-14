"""
        If you add a new motorway in the list above,make sure you add another nested if else statement below
       """
import myFunctions
from openpyxl import load_workbook
import logging

def load(file_path,m,add,day,mYYYY):
    if(m=='BAC'):
        logger=myFunctions.setup_logger(m,add+m)
        #for first month of day load the template
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/BAC.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:
            #for rest of the days just load the file with prefix month and year eg:MAR2018,not the template
            wb=load_workbook(filename=file_path+'File/BAC/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='CCM'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/CCM.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/CCM/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='EHML'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/EHML.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/EHML/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='ELK'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/ELK.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/ELK/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='LCT'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/LCT.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/LCT/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='M1'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/M1.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/M1/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='M4WCX'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/M4WCX.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/M4WCX/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='M7'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/M7.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/M7/'+m+'_'+mYYYY+'.xlsx')            
    elif(m=='MCL'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/MCL.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/MCL/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='QML'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/QML.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/QML/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='SAB'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/SAB.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/SAB/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='SAT'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/SAT.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/SAT/'+m+'_'+mYYYY+'.xlsx')
    elif(m=='SHB'):
        logger=myFunctions.setup_logger(m,add+m)
        #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
        if(day=='01'):
            wb=load_workbook(filename=file_path+'template/SHB.xlsx')
            myFunctions.Delete_Txt(file_path,m)
        else:                  
            wb=load_workbook(filename=file_path+'File/SHB/'+m+'_'+mYYYY+'.xlsx')
    return (wb,logger)
