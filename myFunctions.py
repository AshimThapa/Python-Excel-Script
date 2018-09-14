import openpyxl
import time
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.styles import Color
from openpyxl.cell import Cell
import logging,logging.handlers
import os


formatter=logging.Formatter('\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
created_time=time.strftime('%d%m%y%H%M%S')

def setup_logger(name,log_file,level=logging.INFO):
        if name!= 'email':
                handler=logging.FileHandler(log_file+'_'+created_time+'.log')
                handler.setFormatter(formatter)
        else:
                handler=logging.handlers.SMTPHandler(mailhost='webmail.interlinkroads.com.au',
                                            fromaddr='casit@eway.net.au',
                                            toaddrs=['athapa@interlinkroads.com.au'],
                                            subject='Auto daily toll notice Error !!!',
                                            credentials='casit@eway.net.au',
                                            secure=None)

        logger=logging.getLogger(name)
        logger.setLevel(logging.INFO)
        logger.addHandler(handler)
        return logger

def setStyleCell(cols,cold,rows,rowd,ws):
        for row in ws.iter_cols(cols,cold,rows,rowd):
            for cell in row:
                my_color=openpyxl.styles.colors.Color(rgb='9B9696')
                my_fill=openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=my_color)
                cell.fill=my_fill
        return;
#function to fill total value for all
def fillTotal(cell_value,total_Trips,total_Amount,ws):
        row=[]
        trips=[]
        amount=[]
        for k in range(len(cell_value)):
                if(cell_value[k]!=0 and total_Trips[k]!=0):
                        row.append(cell_value[k])
                        trips.append(total_Trips[k])
                        amount.append(total_Amount[k])

        for i in range(len(row)):
                ws['E'+str(row[i])]=float(trips[i])
                ws['F'+str(row[i])]=float(amount[i])
                ws['E'+str(row[i])].font=Font(color='0000FF')
                ws['F'+str(row[i])].font=Font(color='0000FF')
                ws['E'+str(row[i])].border=Border(bottom=Side(style='double',color='000000'),top=Side(color='000000',style='thin'))
                ws['F'+str(row[i])].border=Border(bottom=Side(style='double',color='000000'),top=Side(color='000000',style='thin'))
        return


#function to fill admin fees
def Fill_Fee(cellValue,position,admin_fee,ws,c=False):
    if(c==True):
        ws['H'+str(cellValue)]=float(admin_fee[position][0])
        ws['I'+str(cellValue)]=float(admin_fee[position][1])
        ws['J'+str(cellValue)]=float(admin_fee[position][2])
    else:
        ws['H'+str(cellValue)]=float(admin_fee[position][0])
        ws['I'+str(cellValue)]=float(admin_fee[position][1])
        ws['J'+str(cellValue)]=float(admin_fee[position][2])

#function to fill cell values
def InsertCellValues(r,col,v,ws):
    if(len(v)==5):
        d=r[0]
        e=r[1]
        f=r[2]
        ws[d+col[0]]=float(v[0])
        ws[d+col[0]].font=Font(color='0000FF')
        ws[d+col[0]].alignment=Alignment(horizontal='left')
        ws[e+col[0]]=float(v[1])
        ws[f+col[0]]=float(v[2])
        ws[e+col[1]]=float(v[3])
        ws[e+col[1]].font=Font(color='FF0000')
        ws[f+col[1]]=float(v[4])
        ws[f+col[1]].font=Font(color='FF0000')
        totalTrips=float(v[1])+float(v[3])
        totalAmount=float(v[2])+float(v[4])
    elif(len(v)==4):
        c=r[0]
        d=r[1]
        e=r[2]
        f=r[3]
        ws[c+col[0]]=float(v[0])
        ws[d+col[0]]=float(v[1])
        ws[e+col[0]]=float(v[2])
        ws[f+col[0]]=float(v[3])
        ws[c+col[0]].font=Font(color='0000FF')
        ws[d+col[0]].font=Font(color='0000FF')
        ws[e+col[0]].font=Font(color='0000FF')
        ws[f+col[0]].font=Font(color='0000FF')
        ws[d+col[0]].alignment=Alignment(horizontal='left')
        totalTrips=float(v[2])
        totalAmount=float(v[3])
    elif(len(v)==7):
        c=r[0]
        d=r[1]
        e=r[2]
        f=r[3]
        ws[c+col[0]]=v[0]
        ws[c+col[0]].alignment=Alignment(horizontal='left')
        ws[c+col[0]].font=Font(color='000000',bold=True,size=11)
        ws[e+col[1]]=v[1]
        ws[f+col[1]]=v[2]
        ws[e+col[1]].alignment=Alignment(horizontal='center')
        ws[e+col[1]].font=Font(color='000000',size=11)
        ws[f+col[1]].alignment=Alignment(horizontal='center')
        ws[f+col[1]].font=Font(color='000000',size=11)
        ws[c+col[2]]=float(v[3])
        ws[d+col[2]]=float(v[4])
        ws[d+col[2]].font=Font(color='0000FF')
        ws[d+col[2]].alignment=Alignment(horizontal='left')
        ws[c+col[2]].font=Font(color='0000FF')
        ws[e+col[2]]=float(v[5])
        ws[f+col[2]]=float(v[6])
        ws[e+col[2]].font=Font(color='0000FF')
        ws[f+col[2]].font=Font(color='0000FF')
        totalTrips=float(v[5])
        totalAmount=float(v[6])
    elif(len(v)==6):
        c=r[0]
        d=r[1]
        e=r[2]
        f=r[3]
        ws[c+col[0]]=float(v[0])       
        ws[c+col[0]].font=Font(color='0000FF')        
        ws[d+col[0]]=float(v[1])
        ws[d+col[0]].font=Font(color='0000FF')
        ws[d+col[0]].alignment=Alignment(horizontal='left')
        ws[e+col[0]]=float(v[2])
        ws[e+col[0]].font=Font(color='0000FF')
        ws[f+col[0]]=float(v[3])
        ws[f+col[0]].font=Font(color='0000FF')
        ws[e+col[1]]=float(v[4])
        ws[e+col[1]].font=Font(color='FF0000')
        ws[f+col[1]]=float(v[5])
        ws[f+col[1]].font=Font(color='FF0000')
        totalTrips=float(v[2])+float(v[4])
        totalAmount=float(v[3])+float(v[5])        
    elif(len(v)==9):
        c=r[0]
        d=r[1]
        e=r[2]
        f=r[3]
        ws[c+col[0]]=v[0]
        ws[c+col[0]].alignment=Alignment(horizontal='left')
        ws[c+col[0]].font=Font(color='000000',bold=True,size=11)
        ws[e+col[1]]=v[1]
        ws[f+col[1]]=v[2]
        ws[e+col[1]].alignment=Alignment(horizontal='center')
        ws[e+col[1]].font=Font(color='000000',size=11)
        ws[f+col[1]].alignment=Alignment(horizontal='center')
        ws[f+col[1]].font=Font(color='000000',size=11)
        ws[c+col[2]]=float(v[3])
        ws[c+col[2]].font=Font(color='0000FF')
        ws[d+col[2]]=float(v[4])
        ws[d+col[2]].font=Font(color='0000FF')
        ws[d+col[2]].alignment=Alignment(horizontal='left')
        ws[e+col[2]]=float(v[5])
        ws[f+col[2]]=float(v[6])
        ws[e+col[3]]=float(v[7])
        ws[e+col[3]].font=Font(color='FF0000')
        ws[f+col[3]]=float(v[8])
        ws[f+col[3]].font=Font(color='FF0000')
        totalTrips=float(v[5])+float(v[7])
        totalAmount=float(v[6])+float(v[8])
    return (totalTrips,totalAmount)

#function to substring filename
def getFilenameDetails(filename):
    file_detail=filename[0:8]        
    receiverId=filename[3:6]
    file_type=filename[6:8]
    file_no=filename[8:]
    return (receiverId,file_type,file_no,file_detail)

#function to determine filling postion
def startFillingFrom(file_07_111,file_07_108,file_33_111,file_33_108):
    if(file_07_111==1):
        last_index_07108=24
    elif(file_07_111==2):
        last_index_07108=26
    elif(file_07_111==3):
        last_index_07108=28
    elif(file_07_111==0):
        last_index_07108=20
        
    if(file_07_108==1):
        next_index_33111=last_index_07108+6
    elif(file_07_108==2):
        next_index_33111=last_index_07108+8
    elif(file_07_108==3):
        next_index_33111=last_index_07108+10
    elif(file_07_108==0):
        next_index_33111=last_index_07108

    if(file_33_111==1):
        next_index_33108=next_index_33111+4
    elif(file_33_111==2):
        next_index_33108=next_index_33111+5
    elif(file_33_111==3):
        next_index_33108=next_index_33111+6
    elif(file_33_111==0):
        next_index_33108=next_index_33111

    return (last_index_07108,next_index_33111,next_index_33108)

#function to fill admin fee section text
def Fill_Fee_Text(index,ws,f_list,f_detail):
        ws['F'+str(index)]=int(f_detail)
        ws['F'+str(index)].font=Font(color='0000FF')
        ws['G'+str(index)]=int(f_list)
        ws['G'+str(index)].font=Font(color='0000FF')
        ws['G'+str(index)].alignment=Alignment(horizontal='left')
        return
def Delete_Txt(file_path,m):
        path=file_path+"/Email/"+m+".txt"
        if os.path.isfile(path):
                          os.remove(path)
        return

