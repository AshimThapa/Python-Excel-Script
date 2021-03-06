"""
Written by Ashim Thapa 27/02/2018

It generates Auto Daily Toll Notice for 13 motorways.

This python script loads an excel file as template
Reads the text file values and
The values are entered into the excel sheet.
"""
import EmailLib
import openpyxl
import time
import sys
import myFunctions
import loadTemplate
import MainPage
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.styles import Color
from openpyxl.cell import Cell
import logging
from calendar import monthrange
from pathlib import Path

#path where your files and log files are located
add='C:......../Logs/'
file_path='C:...../'

lines=[]

#Column list where data goes
r=['D','E','F']
r1=['C','D','E','F']

#GET TODAY'S DAY IN DAY NUMBER AND MONTH CHAR FORMAT (02MAR)
namesheet=time.strftime("%d%b")
#GET TODAY'S DAY IN MYYYY FORMAT
mYYYY=time.strftime("%b%Y")
#GET TODAY'S DAY IN NUMBER
day=time.strftime("%d")




month=int(time.strftime('%m'))
year=int(time.strftime('%Y'))
last_day=monthrange(year,month)[1]
#Motorway Short Descriptions

"""
For future add your motorways here and make sure template files are ready for that motorway
"""
motorway=['BAC','CCM','EHML','ELK','LCT','M1','M4WCX','M7','MCL','QML','SAB','SAT','SHB']
#motorway=['CCM']
path_list=[]






#Loop through each MOTORWAY 
for m in motorway:
#Define LOGS
    #log_name=add+m+'.txt'
    #logging.basicConfig(filename=add+m+'.txt',level=logging.DEBUG,format='\n\n%(asctime)s %(message)s',datefmt='%H:%M:%S')
    #logger=logging.getLogger(m)

#boolean value to trigger for more than one file
    error_flag=False
#list that holds admin fees
    admin_fee_33111=[]
    admin_fee_33108=[]
#list to hold file names
    f_list_33111=[]
    f_list_33108=[]

#Counters
    c_111=0
    c_108=0
    c_33_111=0
    c_33_108=0
#cell value i.e ROW number/index
    last_index_07108=0
    next_index_33111=0
    next_index_33108=0
    admin_next=0
#number of file counter
    file_07_111=0
    file_07_108=0
    file_33_111=0
    file_33_108=0
#total of each individual file
    tTrips=0
    tAmount=0
    total_index=0
#Total for each file type
    total_07111_Trip=0
    total_07111_Amount=0
    total_07108_Trip=0
    total_07108_Amount=0
    total_33111_Trip=0
    total_33111_Amount=0
    total_33108_Trip=0
    total_33108_Amount=0

   
#indexing variable
    new_value=0

    try:
        #LOAD THE RIGHT TEMPLATE
        print(m)        
        (wb,logger)=loadTemplate.load(file_path,m,add,day,mYYYY)
        logger.info("start of Process\n")
        logger.info("=======================================================================\n")
        logger.info("Template Loaded\n")
   
        
       #NOW CHECK IF SHEET ALREADY EXIST for today's date
        try:
            if not namesheet in wb.sheetnames:            
                            
                logger.info("========================================================================================\n")
                wb.copy_worksheet(wb['temp'])
                ws=wb['temp Copy']
                ws.title=namesheet
                logger.info("template copied\n")
                        
                #Write some graphics
                img=openpyxl.drawing.image.Image(file_path+'images/Eway1.png')
                ws.add_image(img,'C5')
                img=openpyxl.drawing.image.Image(file_path+'images/interlink2.png')
                ws.add_image(img,'I5')
                
                #call funtion to fill pattern
                myFunctions.setStyleCell(1,1,1,83,ws)
                myFunctions.setStyleCell(1,14,1,3,ws)
                myFunctions.setStyleCell(12,14,1,83,ws)
                myFunctions.setStyleCell(2,13,83,83,ws)
                logger.info("Images loaded and background color set\n")

            #NOW READ DATA FILE
                logger.info("Reading Data from text file\n")
                with open(file_path+'data/'+m+'.txt','r') as data:
                    lines=data.readlines()
            #get the total number of files for each reciver for a type 33 or 07
                for i in range(len(lines)):
                    try:
                        data=lines[i].split(",")
                        #call function to substring file no,file type,receiver
                        (rId,f_t,f_no,f_detail)=myFunctions.getFilenameDetails(data[0])
                        
                        if(f_t=='07'):
                            if(rId=='111'):
                                file_07_111+=1
                            elif(rId=='108'):
                                file_07_108+=1
                        elif(f_t=='33'):
                            if(rId=='111'):
                                file_33_111+=1
                            elif(rId=='108'):
                                file_33_108+=1
                        
                    except Exception as err:
                        logger.exception(str(err))
                logger.info("************************************************************************************\n")
                logger.info("Data files read, no of file to be written to excell\n")
                logger.info("File_07_111: "+str(file_07_111)+"\n")
                logger.info("File_07_108: "+str(file_07_108)+"\n")
                logger.info("File_33_111: "+str(file_33_111)+"\n")
                logger.info("File_33_108: "+str(file_33_108)+"\n")
                
                logger.info("*************************************************************************************\n")

#creates a text file to alert at the month end that there were more than one files on a day
                if(file_07_111 > 1 or file_07_108 > 1 or file_33_111 >1 or file_33_108 > 1):
                    with open(file_path+"/Email/"+m+".txt","a") as f:
                        f.write("More than one trip file received on date: "+namesheet+"\n")

                for index in range(len(lines)):
            #Read and process first file
                    try:
                        rawData=lines[index].split(",")
                #substring to get filetype and receiverid
                        (rId,f_t,f_no,f_detail)=myFunctions.getFilenameDetails(rawData[0])
                        
                        #Call function to get index value of excel cells
                        (last_index_07108,next_index_33111,next_index_33108)=myFunctions.startFillingFrom(file_07_111,file_07_108,file_33_111,file_33_108)

                        #process files
                        if(f_t=='07' and rId=='111'):
                            logger.info("Start processing File 07_111\n")
                        
                            if(c_111==0):
                                ws['C20']='Tag Transaction File - M5 - File 111'
                                ws['C20'].font=Font(color='000000',bold=True,size=11)
                                ws['C20'].alignment=Alignment(horizontal='left')
                                ws['C22']=int(f_detail)
                                ws['C23']='Rejected Trips'
                                ws['C23'].font=Font(color='ff0000')
                                ws['E21']='Trips'
                                ws['F21']='Amount'
                                ws['E21'].alignment=Alignment(horizontal='center')
                                ws['E21'].font=Font(color='000000',size=11)
                                ws['F21'].alignment=Alignment(horizontal='center')
                                ws['F21'].font=Font(color='000000',size=11)
                                
                                c=['22','23']
                                v=[f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r,c,v,ws)
                                total_07111_Trip+=tT
                                total_07111_Amount+=tA
                                c_111 +=1
                                
                            elif(c_111==1):
                                #more than one file
                                c=['24','25']
                                ws['C25']='Rejected Trips'
                                ws['C25'].font=Font(color='ff0000')
                                v=[f_detail,f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_07111_Trip+=tT
                                total_07111_Amount+=tA
                                c_111+=1
                                
                            elif(c_111==2):
                                c=['26','27']
                                ws['C27']='Rejected Trips'
                                ws['C27'].font=Font(color='ff0000')
                                v=[f_detail,f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_07111_Trip+=tT
                                total_07111_Amount+=tA
                                c_111+=1
                                
                            elif(c_111>=3):
                                logger.error('ERROR!!! Number of files to be processed more than 3, System cannot process the file,process manually\n')
                                sys.exit()

                            logger.info('File 07_111 processed\n')
                            logger.info('=======================================================================\n')                        
                        
                #now fill the data for file 07 from 108
                        elif(f_t=='07' and rId=='108'):
                            logger.info('Start processing File 07_108\n')
                            #call function
                            (last_index_07108,next_index_33111,next_index_33108)=myFunctions.startFillingFrom(file_07_111,file_07_108,file_33_111,file_33_108)
                            
                            if(c_108==0):
                                c0=str(last_index_07108+2)
                                c1=str(last_index_07108+3)
                                c2=str(last_index_07108+4)
                                c3=str(last_index_07108+5)
                                c=[c0,c1,c2,c3]
                                ws['C'+c3]='Rejected Trips'
                                ws['C'+c3].font=Font(color='ff0000')
                                v=['Tag Transaction File -M5 - FILE 108','Trips','Amount',f_detail,f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_07108_Trip+=tT
                                total_07108_Amount+=tA
                                c_108+=1
                                
                            elif(c_108==1):
                                c0=str(last_index_07108+6)
                                c1=str(last_index_07108+7)
                                c=[c0,c1]
                                ws['C'+c1]='Rejected Trips'
                                ws['C'+c1].font=Font(color='ff0000')
                                v=[f_detail,f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_07108_Trip+=tT
                                total_07108_Amount+=tA
                                c_108+=1
                                
                            elif(c_108==2):
                                c0=str(last_index_07108+8)
                                c1=str(last_index_07108+9)
                                c=[c0,c1]
                                ws['C'+c1]='Rejected Trips'
                                ws['C'+c1].font=Font(color='ff0000')
                                v=[f_detail,f_no,rawData[1],rawData[2],rawData[3],rawData[4]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_07108_Trip+=tT
                                total_07108_Amount+=tA
                                c_108+=1
                                
                            elif(c_108>=3):
                                logger.error("ERROR!!! Number of files to be processed more than 3, System cannot process the file,check manually\n")
                                
                            logger.info("File 07_108 processed\n")
                            logger.info("=======================================================================\n")
                            
                            
                            
                        
                #now fill the data for file 33 from 111
                        elif(f_t=='33' and rId=='111'):
                            logger.info("Start processing File 33_111\n")
                            
                            total_index=0
                            (last_index_07108,next_index_33111,next_index_33108)=myFunctions.startFillingFrom(file_07_111,file_07_108,file_33_111,file_33_108)
                            
                            if(c_33_111==0):
                                f_list_33111.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33111=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33111.append(list_33111)
                                
                                c0=str(next_index_33111+2)
                                c1=str(next_index_33111+3)
                                c2=str(next_index_33111+4)
                                c=[c0,c1,c2]
                                v=['Manual Debit -M5 - FILE 111','Trips','Amount',f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33111_Trip+=tT
                                total_33111_Amount+=tA
                                c_33_111+=1
                                
                            elif(c_33_111==1):
                                f_list_33111.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33111=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33111.append(list_33111)
                                c0=str(next_index_33111+5)
                                c=[c0]
                                v=[f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33111_Trip+=tT
                                total_33111_Amount+=tA
                                c_33_111+=1
                                
                            elif(c_33_111==2):
                                f_list_33111.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33111=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33111.append(list_33111)
                                c0=str(next_index_33111+6)
                                c=[c0]
                                v=[f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33111_Trip+=tT
                                total_33111_Amount+=tA
                                c_33_111+=1
                                
                            elif(c_33_111>=3):
                                logger.error("ERROR!!! Number of files to be processed more than 3, System cannot process the file,check manually\n")
                                
                            logger.info("File 33_111 processed\n")
                            logger.info("=======================================================================\n")
                            
                                      
                #now fill the data for file 33 from 108
                        elif(f_t=='33' and rId=='108'):
                            logger.info("Start processing File 33_108\n")
                            total_index=0
                            (last_index_07108,next_index_33111,next_index_33108)=myFunctions.startFillingFrom(file_07_111,file_07_108,file_33_111,file_33_108)

                            if(c_33_108==0):
                                f_list_33108.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33108=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33108.append(list_33108)
                                c0=str(next_index_33108+3)
                                c1=str(next_index_33108+4)
                                c2=str(next_index_33108+5)
                                total_index=str(next_index_33108+6)
                                c=[c0,c1,c2]
                                v=['Manual Debit -M5 - FILE 108','Trips','Amount',f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33108_Trip+=tT
                                total_33108_Amount+=tA
                                c_33_108+=1
                                
                            elif(c_33_108==1):
                                f_list_33108.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33108=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33108.append(list_33108)
                                c0=str(next_index_33108+6)
                                total_index=str(next_index_33108+7)
                                c=[c0]
                                v=[f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33108_Trip+=tT
                                total_33108_Amount+=tA
                                c_33_108+=1
                                
                            elif(c_33_108==2):
                                f_list_33108.append(f_no)
                                if(m not in ('ELK','QML','MCL')):
                                    list_33108=[rawData[3],rawData[4],rawData[5]]
                                    admin_fee_33108.append(list_33108)
                                c0=str(next_index_33108+7)
                                total_index=str(next_index_33108+8)
                                c=[c0]
                                v=[f_detail,f_no,rawData[1],rawData[2]]
                                (tT,tA)=myFunctions.InsertCellValues(r1,c,v,ws)
                                total_33108_Trip+=tT
                                total_33108_Amount+=tA
                                c_33_108+=1
                                
                            elif(c_33_108>=3):
                                logger.error("ERROR!!! Number of files to be processed more than 3, System cannot process the file,check manually\n")
                            logger.info("File 33_108 processed\n")
                            
                            
                    except Exception as err:
                            logger.exception(str(err))
                logger.info("=======================================================================\n")             
                logger.info("This section is where formula is being processed\n")       
                #calcute total for trips and amount for each file type from 111 and 108
                #If you need to update formulas update it over here
                total1=[total_07111_Trip,total_07108_Trip,total_33111_Trip,total_33108_Trip]
                total2=[total_07111_Amount,total_07108_Amount,total_33111_Amount,total_33108_Amount]
                
                if(next_index_33108!=0):
                    new_value=next_index_33108+1
                else:
                    new_value=0
                ref=[last_index_07108,next_index_33111,new_value,total_index]

                #The total values calcualtion
                ws['J21']=round((total_07111_Amount/11)*10,2)
                if total_07108_Amount!=0:
                    ws['J27']=round((total_07108_Amount/11)*10,2)
                    
                if(m=='BAC'):
                    ws['J22']=round((-total_07111_Trip*x),2)
                    if total_07108_Trip!=0:
                        ws['J28']=round((-total_07108_Trip*x),2)
                elif(m=='EHML'):
                    ws['J22']=round((-total_07111_Trip*x),2)
                    ws['J28']=round((-total_07108_Trip*x),2)
                    ws['J34']=round((-total_33111_Trip*x),2)
                    ws['J40']=round((-total_33108_Trip*x6),2)
                elif(m in ('LCT','M1','M4WCX')):
                    ws['J22']=round((-total_07111_Trip*x-total_07111_Amount*x/100),2)
                    ws['J28']=round((-total_07108_Trip*x-total_07108_Amount*x/100),2)
                    ws['J34']=round((-total_33111_Trip*x-total_33111_Amount*x/100),2)
                    ws['J40']=round((-total_33108_Trip*x-total_33108_Amount*x/100),2)
                elif(m=='QML'):
                    ws['J22']=x
                    ws['J28']=x
                    ws['J34']=x
                    ws['J40']=x
                elif(m=='SAB'):
                    ws['J22']=round((-total_07111_Trip*x-(float(ws['J21'].value)*x)/100),2)
                    ws['J28']=round((-total_07108_Trip*x-(float(ws['J27'].value)*x)/100),2)
                elif(m in ('MCL','SAT','SHB')):
                    if(m=='SAT'):
                        ws['J22']=round((-total_07111_Trip*x),2)
                        ws['J28']=round((-total_07108_Trip*x),2)
                    else:
                        ws['J22']=round((-total_07111_Trip*x),2)
                        if(total_07108_Trip!=0):
                            ws['J28']=round((-total_07108_Trip*x),2)
                        ws['J34']=round((-total_33111_Trip*x),2)
                        if(total_33108_Trip!=0):
                            ws['J40']=round((-total_33108_Trip*x),2)                        
                else:
                    ws['J22']=round((-total_07111_Trip*x-total_07111_Amount*x/100),2)
                    if(total_07108_Trip!=0):
                        ws['J28']=round((-total_07108_Trip*x-total_07108_Amount*x/100),2)
                    
                    ws['J34']=round((-total_33111_Trip*x-total_33111_Amount*x/100),2)
                    if(total_33108_Trip!=0):
                        ws['J40']=round((-total_33108_Trip*x-total_33108_Amount*x/100),2)
                    
                    
                

                if(m not in ('BAC','SAB','SAT')):
                   ws['J33']=round((total_33111_Amount/11)*10,2)
                   if(total_33108_Amount!=0):
                       ws['J39']=round((total_33108_Amount/11)*10,2)                  
               
                    

                #call function to fill totals
                myFunctions.fillTotal(ref,total1,total2,ws)
                logger.info("All total has been calculated and filled in the cells\n")
                logger.info("Now processing LPN fee, Letter1 fee and Letter 2 fee\n")

                #total processing fees for 33 files

                if m not in ('ELK','MCL','QML'):
                    if(file_33_111!=0):
                                
                        indx=54
                        ws['C'+str(indx)]="Manual Debit Administration Fee-111"
                        ws['C'+str(indx)].font=Font(color='000000',bold=True,size=11)
                        ws['C'+str(indx)].alignment=Alignment(horizontal='left')
                        ws['F'+str(indx)]=int(f_detail)
                        ws['F'+str(indx)].font=Font(color='0000FF')
                        ws['G'+str(indx)]=int(f_list_33111[0])
                        ws['G'+str(indx)].font=Font(color='0000FF')
                        ws['G'+str(indx)].alignment=Alignment(horizontal='left')
                        
                #call function to fill admin fee values
                        myFunctions.Fill_Fee(indx,0,admin_fee_33111,ws,True)
                        indx+=1
                        admin_next=indx
                        temp_count=file_33_111-1
                        if(temp_count==1):
                            myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33111[1],f_detail)
                            myFunctions.Fill_Fee(indx,1,admin_fee_33111,ws,True)
                            indx+=1
                            admin_next=indx            

                        if(temp_count==2):
                            myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33111[1],f_detail)
                            myFunctions.Fill_Fee(indx,2,admin_fee_33111,ws,True)
                            indx+=1
                            admin_next=indx
                            temp_count-=1
                            if(temp_count==1):
                                myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33111[2],f_detail)
                                myFunctions.Fill_Fee(indx,1,admin_fee_33111,ws,True)
                                indx+=1
                                admin_next=indx              

                    if(file_33_108!=0):
                        ws['C'+str(admin_next)]="Manual Debit Administration Fee-108"
                        ws['C'+str(admin_next)].font=Font(color='000000',bold=True,size=11)
                        ws['C'+str(admin_next)].alignment=Alignment(horizontal='left')
                        ws['F'+str(admin_next)]=int(f_detail)
                        ws['F'+str(admin_next)].font=Font(color='0000FF')
                        ws['G'+str(admin_next)]=int(f_list_33108[0])
                        ws['G'+str(admin_next)].font=Font(color='0000FF')
                        ws['G'+str(admin_next)].alignment=Alignment(horizontal='left')
                        myFunctions.Fill_Fee(indx,0,admin_fee_33108,ws)
                        indx+=1
                        admin_next=indx
                        temp_count=file_33_108-1
                        if(temp_count==1):
                            myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33108[1],f_detail)
                            myFunctions.Fill_Fee(indx,1,admin_fee_33108,ws)
                            indx+=1
                            admin_next=indx
                            
                        if(temp_count==2):
                            myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33108[1],f_detail)
                            myFunctions.Fill_Fee(indx,2,admin_fee_33108,ws)
                            indx+=1
                            admin_next=indx
                            temp_count-=1
                            if(temp_count==1):
                                myFunctions.Fill_Fee_Text(admin_next,ws,f_list_33108[2],f_detail)
                                myFunctions.Fill_Fee(indx,2,admin_fee_33108,ws)
                                indx+=1
                                admin_next=indx
                logger.info("Data inserted into file\n")
                logger.info("Last_day "+str(last_day))
                logger.info("today: "+str(day))
                logger.info("Validating last day of month")
                if(int(day)==int(last_day)):
                    logger.info("Creating month end data")
                    logger.info("checking text file for multiple file entries")
                    #if file exist enter path to a list
                    if Path(file_path+"/Email/"+m+".txt").is_file():
                        path_list.append(file_path+"/Email/"+m+".txt")
                    #call funcion to fill main page                     
                    MainPage.CallMainPage(m,wb,logger) #,file_07_111,file_07_108,file_33_111,file_33_108)
                logger.info("Excell file Created and saved.logger file Created as well\n")
                wb.save(file_path+'File/'+m+'/'+m+'_'+mYYYY+'.xlsx')
                logger.info("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n")    
                logger.info("End of Process")
            else:
                raise FileExistsError("Sheet already exist with current date as name,please check the excel file\n")
                log.info("Sheet already exist with current date as name,please check the excel file")
                log=myFunctions.setup_logger('email','test.log')
                log.exception("Sheet Already exist with that name")
            
                
        except Exception as err:
            logger.exception(str(err))             
            
    except Exception as err:
        logger.exception(str(err))
        log=myFunctions.setup_logger('email','test.log')
        log.exception("File not found,please have a look, thanks CAS IT  ")
 #if last day of month send ALERT EMAIL for MAIN PAGE     
if(int(day)==int(last_day)):
     if not path_list:
         logger.info('No Multiple trips for this month for any motorways')
     else:
         EmailLib.sendEmail(path_list,file_path)

     

 
